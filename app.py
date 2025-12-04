import io
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

TEMPLATE_FILE = "External Examination Marksheet_blank.xlsx"


def clean_enrollment(enroll_str):
    if pd.isna(enroll_str):
        return None
    s = str(enroll_str).strip()
    if "e" in s.lower():
        try:
            return str(int(float(s)))
        except Exception:
            return s
    return s


def get_data_rows_in_sheet(ws):
    # Row 7 header, 8–37 = 1–30, 38 header, 39–53 = 31–45
    rows = []
    for sr in range(1, 31):
        rows.append((sr, 7 + sr))      # 8–37
    for sr in range(31, 46):
        rows.append((sr, 8 + sr))      # 39–53
    return rows


def build_marksheet(student_df, batch_size: int, branch_wise: bool) -> bytes:
    # 1) clean + remove duplicate enrollment
    df = student_df.copy()
    df["Enrollment no"] = df["Enrollment no"].apply(clean_enrollment)
    df = df.drop_duplicates(subset=["Enrollment no"], keep="first")

    # 2) sorting
    if branch_wise and "Branch" in df.columns:
        df = df.sort_values(
            by=["Branch", "Enrollment no"],
            ascending=[True, True]
        ).reset_index(drop=True)
    else:
        df = df.sort_values(
            "Enrollment no",
            key=lambda x: pd.to_numeric(x, errors="coerce")
        ).reset_index(drop=True)

    # 3) batch numbers
    df["Batch"] = (df.index // batch_size) + 1
    max_batch = int(df["Batch"].max())

    # 4) load template workbook
    wb = load_workbook(TEMPLATE_FILE)

    # 5) fill sheets
    for b in range(1, max_batch + 1):
        sheet_name = f"Batch {b}"
        batch_students = df[df["Batch"] == b].reset_index(drop=True)

        if sheet_name not in wb.sheetnames:
            # copy last sheet as template
            template_name = wb.sheetnames[-1]
            template_ws = wb[template_name]
            new_ws = wb.copy_worksheet(template_ws)
            new_ws.title = sheet_name
            ws = new_ws
        else:
            ws = wb[sheet_name]

        valid_rows = get_data_rows_in_sheet(ws)

        for idx, row in batch_students.iterrows():
            if idx >= len(valid_rows):
                break
            sr_no, excel_row = valid_rows[idx]
            ws.cell(row=excel_row, column=1).value = sr_no
            ws.cell(row=excel_row, column=2).value = row["Enrollment no"]
            ws.cell(row=excel_row, column=3).value = row["Name"]

        # clear remaining rows (optional, keep blank but safe)
        for idx in range(len(batch_students), len(valid_rows)):
            _, excel_row = valid_rows[idx]
            ws.cell(row=excel_row, column=1).value = None
            ws.cell(row=excel_row, column=2).value = None
            ws.cell(row=excel_row, column=3).value = None

    # 6) save to memory (no disk)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ----------------- STREAMLIT UI -----------------

st.title("External Viva Marksheet Generator")

uploaded = st.file_uploader("Upload Student-List Excel", type=["xlsx"])
batch_size = st.number_input("Students per batch", min_value=1, max_value=45, value=45)
branch_wise = st.checkbox("Sort branch-wise then enrollment-wise", value=True)

if uploaded is not None:
    try:
        df_students = pd.read_excel(uploaded)
        st.write("Preview of uploaded data:")
        st.dataframe(df_students.head())

        if st.button("Generate Marksheet"):
            output_bytes = build_marksheet(df_students, int(batch_size), branch_wise)

            st.success("Marksheet generated successfully!")
            st.download_button(
                label="Download Excel Marksheet",
                data=output_bytes,
                file_name="External-Examination-Marksheet_FINAL.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception as e:
        st.error(f"Error reading file: {e}")

